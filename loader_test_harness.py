"""
Created on May 10, 2017
@author: richard.thomas
"""

import sys
import logging
import re
from openpyxl import load_workbook


class LoaderTestHarness:

    def __init__(self, xlsx_path):
        self.xlsxPath = xlsx_path
        self.test_case = None
        self.section = None
        self.tabs = []
        self.fields = []
        self.properties = []
        self.xml_path = None
        self.wb = load_workbook(self.xlsxPath)
        self.ws = self.wb.active
        self.cases = []

    @staticmethod
    def property_tokenizer(prop_string):
        section_list = ["Carrier", "FleetEquipment", "Driver"]
        regex = re.compile("(?=({0}))".format("|".join(map(re.escape, section_list))))

        this_property = prop_string.split('=')
        operation = this_property[1]
        this_property = this_property[0].split(':')
        condition = this_property[0].strip()
        level = this_property[1]
        section_tab_field = this_property[2]
        section = re.findall(regex, section_tab_field)[0]
        if level == 'tab':
            tab = section_tab_field[len(section):]
            field = None
        elif level == 'field':
            tab = None
            field = section_tab_field[len(section):]
        else:
            tab = None
            field = None
        return Property(prop_string, condition, level, section, tab, field, operation)

    def base_prop_split(prop_string):
        prop = prop_string.split("=")
        return Property(prop_string, None, None, None, None, None, None, prop[0], prop[1])

    def load_test_data(self):
        for index, row in enumerate(self.ws.iter_rows()):
            if index > 1:
                case = TestCase()
                case.number = str(row[0].value)
                if case.number is None:
                    logging.error("Test Case number is required")
                case.sections = row[1].value
                if case.sections is None:
                    logging.error("Section is required")
                if case.sections is not None:
                    case.sections = [section.strip() for section in case.sections.split(",")]
                case.tabs = row[2].value
                if case.tabs is not None:
                    case.tabs = [tab.strip() for tab in case.tabs.split(",")]
                case.fields = row[3].value
                if case.fields is not None:
                    case.fields = [field.strip() for field in case.fields.split(",")]
                if (case.tabs is None) and (case.fields is None):
                    logging.error("Please include either tabs or fields")
                props_list = row[4].value
                if props_list is None:
                    logging.error("Please include desired properties")
                else:
                    case.properties = []
                    for prop in props_list.split(','):
                        case.properties.append(LoaderTestHarness.property_tokenizer(prop))
                case.xml_path = "./" + row[5].value
                if case.xml_path is None:
                    logging.error("Please specify path of XML file to be loaded")
                self.cases.append(case)
        return self.cases

    def base_properties(self):
        base_props = []
        props_list = self.ws['E2'].value
        for prop in props_list.split(','):
            base_props.append(LoaderTestHarness.base_prop_split(prop))
        return base_props

class Property:

    def __init__(self, prop_string=None, condition=None, level=None, section=None, tab=None, field=None,
                 operation=None, property=None, value=None):
        self.prop_string = prop_string
        self.condition = condition
        self.level = level
        self.section = section
        self.tab = tab
        self.field = field
        self.operation = operation


def main():
    loader = LoaderTestHarness(sys.argv[1])
    loader.load_test_data()

if __name__ == '__main__':
    main()


class TestCase:

    def __init__(self):
        self.number = None
        self.sections = None
        self.tabs = []
        self.fields = []
        self.properties = []
        self.xml_path = None
